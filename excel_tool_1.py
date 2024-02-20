# -*- coding:utf-8 -*-
import openpyxl
import time
def cancel_merge_first_value(file_path, sh_nam):
    """ 取消合并单元格,并且填充空白的单元格 """
    wb = openpyxl.load_workbook(filename=file_path)
    sheets = wb.sheetnames  # 获取所有的表格
    start_index_list = []

    sheet_name_dict = {
        sh_nam: start_index_list
        }


    for sheet in sheets:
        if sheet not in sheet_name_dict.keys():
            continue
        ws = wb[sheet]
        merge_list = ws.merged_cells
        all_merge_index = []
        for merge_area in merge_list:
            index_dict = {'all_key': merge_area.coord, 'max_row': merge_area.max_row, 'min_row': merge_area.min_row}
            start_index_list.append(index_dict)
            min_row, max_row, min_col, max_col = merge_area.min_row, merge_area.max_row, merge_area.min_col, merge_area.max_col
            all_merge_index.append((min_row, max_row, min_col, max_col))
        for merge_index in all_merge_index:
            ws.unmerge_cells(start_row=merge_index[0], end_row=merge_index[1], start_column=merge_index[2], end_column=merge_index[3])
        for use_index in start_index_list:
            all_index = use_index['all_key'].split(':')
            min_index = use_index['min_row']
            max_index = use_index['max_row']
            col_name = all_index[0].split(str(min_index))[0]
            start_value = ws[f'{col_name}{min_index}'].value
            for num_index in range(min_index, max_index+1):
                ws[f'{col_name}{num_index}'] = start_value
    wb.save(file_path)


if __name__ == '__main__':
    #use_file_path = r"C:\yy08.vendor\桌面\tmp_folder\test_merge_excel_cel.xlsx"
    use_file_path = r"C:\yy08.vendor\桌面\tmp_folder\test_merge_excel_cel_tmp_1.xlsx"
    sheet_name_list = [
        "4_board_transb_0_Cache_L1_5",
        "4_board_transb_1_Cache_L1_5",
        "4_board_transb_0_No_Cache",
        "4_board_transb_1_No_Cache",
        "8_board_transb_1_Cache_L1_5",
        "8_board_transb_0_No_Cache",
        "8_board_transb_1_No_Cache"
        ]

    for sheet_name in sheet_name_list:
        cancel_merge_first_value(use_file_path, sheet_name)
